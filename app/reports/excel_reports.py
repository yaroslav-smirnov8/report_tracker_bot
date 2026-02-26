from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from app.core.db import Session
from app.core.models import ChatMember, DailyRecord, Fine, Chat


def get_member_display_name(member):
    return member.full_name or member.user_name or str(member.user_id)


def resolve_report_chat_id(update):
    source_chat = update.message.chat if update.message else update.callback_query.message.chat
    if source_chat.type != "private":
        return source_chat.id, 0
    session = Session()
    chats = session.query(Chat).filter(Chat.id < 0).all()
    session.close()
    if len(chats) == 1:
        return chats[0].id, 1
    return None, len(chats)


def parse_record_date(value):
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except Exception:
        return None


def resolve_course_start_date(session, chat_id, records):
    chat = session.query(Chat).filter_by(id=chat_id).first()
    if chat and chat.start_date:
        try:
            return datetime.strptime(chat.start_date, '%Y-%m-%d').date()
        except Exception:
            return None
    dates = [parse_record_date(record.date) for record in records]
    valid_dates = [value for value in dates if value]
    return min(valid_dates) if valid_dates else None


def expected_report_count(start_date, end_date):
    if not start_date or start_date > end_date:
        return 0
    current_date = start_date
    total = 0
    while current_date <= end_date:
        if current_date.weekday() == 6:
            total += 1
        else:
            total += 2
        current_date += timedelta(days=1)
    return total


def completed_report_count(records):
    completed = 0
    for record in records:
        record_date = parse_record_date(record.date)
        if not record_date:
            continue
        if record_date.weekday() == 6:
            if record.week_hashtag == "1":
                completed += 1
        else:
            if record.morning_hashtag == "1":
                completed += 1
            if record.evening_hashtag == "1":
                completed += 1
    return completed


def style_header(ws, header_font, thin_border, header_fill=None, alignment=None):
    for cell in ws[1]:
        cell.font = header_font
        cell.border = thin_border
        if header_fill:
            cell.fill = header_fill
        if alignment:
            cell.alignment = alignment


def build_date_range(start_date, end_date):
    if not start_date or not end_date or start_date > end_date:
        return []
    current_date = start_date
    dates = []
    while current_date <= end_date:
        dates.append(current_date)
        current_date += timedelta(days=1)
    return dates


def build_record_index(records):
    record_by_date = {}
    for record in records:
        record_date = parse_record_date(record.date)
        if record_date:
            record_by_date[record_date] = record
    return record_by_date


def day_completion(record, day):
    if day.weekday() == 6:
        expected = 1
        completed = 1 if record and record.week_hashtag == "1" else 0
        return expected, completed
    morning_done = record and record.morning_hashtag == "1"
    evening_done = record and record.evening_hashtag == "1"
    expected = 2
    completed = int(bool(morning_done)) + int(bool(evening_done))
    return expected, completed


def calculate_streaks(date_list, record_by_date):
    current_streak = 0
    for day in reversed(date_list):
        expected, completed = day_completion(record_by_date.get(day), day)
        if expected == completed:
            current_streak += 1
        else:
            break
    best_streak = 0
    running = 0
    for day in date_list:
        expected, completed = day_completion(record_by_date.get(day), day)
        if expected == completed:
            running += 1
            if running > best_streak:
                best_streak = running
        else:
            running = 0
    return current_streak, best_streak


def calculate_member_metrics(member_records, date_list):
    record_by_date = build_record_index(member_records)
    expected_morning = 0
    expected_evening = 0
    expected_weekly = 0
    completed_morning = 0
    completed_evening = 0
    completed_weekly = 0
    days_complete = 0
    last_report_date = None

    for day in date_list:
        record = record_by_date.get(day)
        if day.weekday() == 6:
            expected_weekly += 1
            if record and record.week_hashtag == "1":
                completed_weekly += 1
                days_complete += 1
                last_report_date = day if not last_report_date or day > last_report_date else last_report_date
        else:
            expected_morning += 1
            expected_evening += 1
            morning_done = record and record.morning_hashtag == "1"
            evening_done = record and record.evening_hashtag == "1"
            if morning_done:
                completed_morning += 1
                last_report_date = day if not last_report_date or day > last_report_date else last_report_date
            if evening_done:
                completed_evening += 1
                last_report_date = day if not last_report_date or day > last_report_date else last_report_date
            if morning_done and evening_done:
                days_complete += 1

    expected_total = expected_morning + expected_evening + expected_weekly
    completed_total = completed_morning + completed_evening + completed_weekly
    completion_rate = completed_total / expected_total if expected_total else 0
    attendance_rate = days_complete / len(date_list) if date_list else 0
    current_streak, best_streak = calculate_streaks(date_list, record_by_date)

    last7_dates = date_list[-7:] if len(date_list) > 7 else date_list
    expected_last7 = 0
    completed_last7 = 0
    for day in last7_dates:
        record = record_by_date.get(day)
        if day.weekday() == 6:
            expected_last7 += 1
            if record and record.week_hashtag == "1":
                completed_last7 += 1
        else:
            expected_last7 += 2
            if record and record.morning_hashtag == "1":
                completed_last7 += 1
            if record and record.evening_hashtag == "1":
                completed_last7 += 1
    last7_rate = completed_last7 / expected_last7 if expected_last7 else 0

    return record_by_date, {
        "expected_total": expected_total,
        "completed_total": completed_total,
        "completion_rate": completion_rate,
        "expected_morning": expected_morning,
        "expected_evening": expected_evening,
        "expected_weekly": expected_weekly,
        "completed_morning": completed_morning,
        "completed_evening": completed_evening,
        "completed_weekly": completed_weekly,
        "attendance_rate": attendance_rate,
        "current_streak": current_streak,
        "best_streak": best_streak,
        "last_report_date": last_report_date,
        "last7_rate": last7_rate,
        "days_complete": days_complete
    }


def apply_rate_fill(cell, green_fill, yellow_fill, red_fill):
    if cell.value is None:
        return
    try:
        rate = float(cell.value)
    except Exception:
        return
    if rate >= 0.85:
        cell.fill = green_fill
    elif rate >= 0.6:
        cell.fill = yellow_fill
    else:
        cell.fill = red_fill


def create_excel_file(chat_id):
    session = Session()

    members = session.query(ChatMember).filter(ChatMember.chat_id == chat_id).order_by(ChatMember.full_name).all()
    records = session.query(DailyRecord).join(ChatMember, ChatMember.id == DailyRecord.chat_member_id).filter(
        ChatMember.chat_id == chat_id).all()
    fines = session.query(Fine).join(ChatMember, ChatMember.id == Fine.chat_member_id).filter(
        ChatMember.chat_id == chat_id).all()

    records_by_member = {}
    fines_by_member = {}
    for member in members:
        records_by_member[member.id] = []
        fines_by_member[member.id] = []

    for record in records:
        records_by_member.setdefault(record.chat_member_id, []).append(record)
    for fine in fines:
        fines_by_member.setdefault(fine.chat_member_id, []).append(fine)

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF')
    green_fill = PatternFill(start_color='00C853', end_color='00C853', fill_type='solid')
    red_fill = PatternFill(start_color='E53935', end_color='E53935', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFB300', end_color='FFB300', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    course_start_date = resolve_course_start_date(session, chat_id, records)
    course_end_date = date.today()
    if not course_start_date:
        course_start_date = course_end_date
    date_list = build_date_range(course_start_date, course_end_date)
    expected_total = expected_report_count(course_start_date, course_end_date)

    member_stats = []
    member_record_maps = {}
    for member in members:
        member_records = records_by_member.get(member.id, [])
        record_by_date, metrics = calculate_member_metrics(member_records, date_list)
        metrics["name"] = get_member_display_name(member)
        metrics["member_id"] = member.id
        member_stats.append(metrics)
        member_record_maps[member.id] = record_by_date

    member_stats_sorted = sorted(member_stats, key=lambda m: m["completion_rate"], reverse=True)
    member_count = len(members)
    group_expected_total = expected_total * member_count
    group_completed_total = sum(m["completed_total"] for m in member_stats)
    group_completion_rate = group_completed_total / group_expected_total if group_expected_total else 0
    avg_completion_rate = (sum(m["completion_rate"] for m in member_stats) / member_count) if member_count else 0
    best_member = member_stats_sorted[0] if member_stats_sorted else None
    top_streak_member = max(member_stats, key=lambda m: m["best_streak"], default=None)
    current_streak_member = max(member_stats, key=lambda m: m["current_streak"], default=None)
    total_fine_amount = sum(fine.fine_amount or 0 for fine in fines)
    total_fine_count = len(fines)

    overview_ws = wb.create_sheet(title="Overview")
    overview_ws.merge_cells("A1:D1")
    overview_ws["A1"] = "Course Overview"
    overview_ws["A1"].font = title_font
    overview_ws["A1"].alignment = left_align
    overview_ws["A3"] = "Key Metrics"
    overview_ws.merge_cells("A3:B3")
    overview_ws["A3"].font = header_font_white
    overview_ws["A3"].fill = header_fill
    overview_ws["A3"].alignment = left_align

    metrics_rows = [
        ("Course period", f"{course_start_date.isoformat()} — {course_end_date.isoformat()}"),
        ("Participants", member_count),
        ("Expected reports", group_expected_total),
        ("Completed reports", group_completed_total),
        ("Completion rate", group_completion_rate),
        ("Average completion rate", avg_completion_rate),
        ("Best completion", f"{best_member['name']} ({best_member['completion_rate']:.0%})" if best_member else "-"),
        ("Top streak", f"{top_streak_member['name']} ({top_streak_member['best_streak']} days)" if top_streak_member else "-"),
        ("Current streak leader", f"{current_streak_member['name']} ({current_streak_member['current_streak']} days)" if current_streak_member else "-"),
        ("Total fines amount", total_fine_amount),
        ("Total fines count", total_fine_count)
    ]
    current_row = 4
    for label, value in metrics_rows:
        overview_ws[f"A{current_row}"] = label
        overview_ws[f"B{current_row}"] = value
        overview_ws[f"A{current_row}"].font = header_font
        overview_ws[f"A{current_row}"].border = thin_border
        overview_ws[f"B{current_row}"].border = thin_border
        overview_ws[f"A{current_row}"].alignment = left_align
        overview_ws[f"B{current_row}"].alignment = left_align
        if label in ("Completion rate", "Average completion rate"):
            overview_ws[f"B{current_row}"].number_format = "0.00%"
            apply_rate_fill(overview_ws[f"B{current_row}"], green_fill, yellow_fill, red_fill)
        current_row += 1
    overview_ws.column_dimensions['A'].width = 28
    overview_ws.column_dimensions['B'].width = 40

    leaderboard_ws = wb.create_sheet(title="Leaderboard")
    leaderboard_headers = [
        "Student",
        "Expected",
        "Completed",
        "Completion %",
        "Morning %",
        "Evening %",
        "Weekly %",
        "Missed",
        "Current Streak",
        "Best Streak",
        "Last Report",
        "Last 7 Days %"
    ]
    leaderboard_ws.append(leaderboard_headers)
    style_header(leaderboard_ws, header_font_white, thin_border, header_fill, center_align)

    for stats in member_stats_sorted:
        morning_rate = stats["completed_morning"] / stats["expected_morning"] if stats["expected_morning"] else 0
        evening_rate = stats["completed_evening"] / stats["expected_evening"] if stats["expected_evening"] else 0
        weekly_rate = stats["completed_weekly"] / stats["expected_weekly"] if stats["expected_weekly"] else 0
        missed = stats["expected_total"] - stats["completed_total"]
        last_report_text = stats["last_report_date"].isoformat() if stats["last_report_date"] else "-"
        leaderboard_ws.append([
            stats["name"],
            stats["expected_total"],
            stats["completed_total"],
            stats["completion_rate"],
            morning_rate,
            evening_rate,
            weekly_rate,
            missed,
            stats["current_streak"],
            stats["best_streak"],
            last_report_text,
            stats["last7_rate"]
        ])

    for row in leaderboard_ws.iter_rows(min_row=2, max_col=12):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align if cell.column > 1 else left_align
    for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'],
                          [24, 10, 10, 14, 12, 12, 12, 10, 14, 12, 14, 14]):
        leaderboard_ws.column_dimensions[col].width = width
    for cell in leaderboard_ws['D'][1:]:
        cell.number_format = "0.00%"
        apply_rate_fill(cell, green_fill, yellow_fill, red_fill)
    for cell in leaderboard_ws['E'][1:]:
        cell.number_format = "0.00%"
        apply_rate_fill(cell, green_fill, yellow_fill, red_fill)
    for cell in leaderboard_ws['F'][1:]:
        cell.number_format = "0.00%"
        apply_rate_fill(cell, green_fill, yellow_fill, red_fill)
    for cell in leaderboard_ws['G'][1:]:
        cell.number_format = "0.00%"
        apply_rate_fill(cell, green_fill, yellow_fill, red_fill)
    for cell in leaderboard_ws['L'][1:]:
        cell.number_format = "0.00%"
        apply_rate_fill(cell, green_fill, yellow_fill, red_fill)
    leaderboard_ws.freeze_panes = "A2"
    if leaderboard_ws.max_row > 1:
        leaderboard_ws.auto_filter.ref = f"A1:L{leaderboard_ws.max_row}"

    trends_ws = wb.create_sheet(title="Trends")
    trends_headers = ["Date", "Expected", "Completed", "Completion %"]
    trends_ws.append(trends_headers)
    style_header(trends_ws, header_font_white, thin_border, header_fill, center_align)

    for day in date_list:
        expected_day = member_count * (1 if day.weekday() == 6 else 2)
        completed_day = 0
        for member in members:
            record = member_record_maps.get(member.id, {}).get(day)
            if day.weekday() == 6:
                if record and record.week_hashtag == "1":
                    completed_day += 1
            else:
                if record and record.morning_hashtag == "1":
                    completed_day += 1
                if record and record.evening_hashtag == "1":
                    completed_day += 1
        completion_rate = completed_day / expected_day if expected_day else 0
        trends_ws.append([day.isoformat(), expected_day, completed_day, completion_rate])

    for row in trends_ws.iter_rows(min_row=2, max_col=4):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
    for cell in trends_ws['D'][1:]:
        cell.number_format = "0.00%"
        apply_rate_fill(cell, green_fill, yellow_fill, red_fill)
    trends_ws.column_dimensions['A'].width = 14
    trends_ws.column_dimensions['B'].width = 12
    trends_ws.column_dimensions['C'].width = 12
    trends_ws.column_dimensions['D'].width = 14

    if trends_ws.max_row > 1:
        trend_chart = LineChart()
        trend_chart.title = "Completion Rate Trend"
        trend_data = Reference(trends_ws, min_col=4, min_row=1, max_row=trends_ws.max_row)
        trend_categories = Reference(trends_ws, min_col=1, min_row=2, max_row=trends_ws.max_row)
        trend_chart.add_data(trend_data, titles_from_data=True)
        trend_chart.set_categories(trend_categories)
        trend_chart.height = 12
        trend_chart.width = 24
        trends_ws.add_chart(trend_chart, "F2")

    heatmap_ws = wb.create_sheet(title="Heatmap")
    heatmap_headers = ["Date"] + [get_member_display_name(member) for member in members]
    heatmap_ws.append(heatmap_headers)
    style_header(heatmap_ws, header_font_white, thin_border, header_fill, center_align)

    for day in date_list:
        row = [day.isoformat()]
        for member in members:
            record = member_record_maps.get(member.id, {}).get(day)
            expected, completed = day_completion(record, day)
            if completed == 0:
                value = ""
            elif completed == expected:
                value = "OK"
            else:
                value = "1/2"
            row.append(value)
        heatmap_ws.append(row)

    for row_idx, day in enumerate(date_list, start=2):
        for col_idx, member in enumerate(members, start=2):
            record = member_record_maps.get(member.id, {}).get(day)
            expected, completed = day_completion(record, day)
            cell = heatmap_ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = center_align
            if completed == 0:
                cell.fill = red_fill
            elif completed == expected:
                cell.fill = green_fill
            else:
                cell.fill = yellow_fill
    for col in range(1, len(members) + 2):
        column_letter = heatmap_ws.cell(row=1, column=col).column_letter
        heatmap_ws.column_dimensions[column_letter].width = 14 if col > 1 else 12
    heatmap_ws.freeze_panes = "B2"

    records_ws = wb.create_sheet(title="Daily Records")
    records_headers = ["Student", "Date", "Morning Report", "Evening Report", "Weekly Report"]
    records_ws.append(records_headers)
    style_header(records_ws, header_font_white, thin_border, header_fill, center_align)

    for member in members:
        member_records = sorted(records_by_member.get(member.id, []), key=lambda r: r.date, reverse=True)
        for record in member_records:
            record_date = parse_record_date(record.date)
            is_sunday = record_date.weekday() == 6 if record_date else False
            records_ws.append([
                get_member_display_name(member),
                record.date,
                "" if is_sunday else record.morning_hashtag,
                "" if is_sunday else record.evening_hashtag,
                record.week_hashtag if is_sunday else ""
            ])
    for row in records_ws.iter_rows(min_row=2, max_col=5):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align if cell.column > 1 else left_align
            if cell.column_letter in ['C', 'D', 'E']:
                if cell.value == '1':
                    cell.fill = green_fill
                elif cell.value == '0':
                    cell.fill = red_fill
    for col in ['A', 'B', 'C', 'D', 'E']:
        records_ws.column_dimensions[col].width = 22
    records_ws.freeze_panes = "A2"
    if records_ws.max_row > 1:
        records_ws.auto_filter.ref = f"A1:E{records_ws.max_row}"

    fines_ws = wb.create_sheet(title="Fines")
    fines_headers = ["Student", "Fine Type", "Fine Amount", "Paid Date"]
    fines_ws.append(fines_headers)
    style_header(fines_ws, header_font_white, thin_border, header_fill, center_align)

    for member in members:
        member_fines = fines_by_member.get(member.id, [])
        for fine in member_fines:
            fines_ws.append([
                get_member_display_name(member),
                fine.report_type,
                fine.fine_amount,
                fine.date_paid
            ])
    for row in fines_ws.iter_rows(min_row=2, max_col=4):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align if cell.column > 1 else left_align
    for col in ['A', 'B', 'C', 'D']:
        fines_ws.column_dimensions[col].width = 20

    if leaderboard_ws.max_row > 1:
        completion_chart = BarChart()
        completion_chart.title = "Completion Rate by Student"
        completion_data = Reference(leaderboard_ws, min_col=4, min_row=1, max_row=leaderboard_ws.max_row)
        completion_categories = Reference(leaderboard_ws, min_col=1, min_row=2, max_row=leaderboard_ws.max_row)
        completion_chart.add_data(completion_data, titles_from_data=True)
        completion_chart.set_categories(completion_categories)
        completion_chart.height = 10
        completion_chart.width = 18
        overview_ws.add_chart(completion_chart, "D3")

    if trends_ws.max_row > 1:
        overview_trend_chart = LineChart()
        overview_trend_chart.title = "Daily Completion Rate"
        overview_trend_data = Reference(trends_ws, min_col=4, min_row=1, max_row=trends_ws.max_row)
        overview_trend_categories = Reference(trends_ws, min_col=1, min_row=2, max_row=trends_ws.max_row)
        overview_trend_chart.add_data(overview_trend_data, titles_from_data=True)
        overview_trend_chart.set_categories(overview_trend_categories)
        overview_trend_chart.height = 10
        overview_trend_chart.width = 18
        overview_ws.add_chart(overview_trend_chart, "D20")

    session.close()
    return wb


def send_excel_file_in_private(update, context):
    chat_id, group_count = resolve_report_chat_id(update)
    if not chat_id:
        target_user_id = update.callback_query.from_user.id if update.callback_query else update.message.chat_id
        if group_count == 0:
            context.bot.send_message(chat_id=target_user_id, text="No group chats found. Use this in the group chat.")
        else:
            context.bot.send_message(chat_id=target_user_id, text="Multiple groups found. Use this in the target group.")
        return False
    workbook = create_excel_file(chat_id)
    file_name = f"chat_{chat_id}_report.xlsx"
    workbook.save(file_name)

    context.bot.send_document(chat_id=update.callback_query.from_user.id if update.callback_query else update.message.chat_id,
                              document=open(file_name, 'rb'))
    return True
